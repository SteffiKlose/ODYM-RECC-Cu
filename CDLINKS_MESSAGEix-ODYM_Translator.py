# -*- coding: utf-8 -*-
"""
Created on Fri Jul 19 11:10:36 2019

@author: sklose
"""

# -*- coding: utf-8 -*-
"""
File CDLINKS_MESSAGEix-ODYM_Translator

Parse, evaluate, and export IAM scenario results for Electricity generation technologies

"""
# Import required libraries:
import os
import sys
import logging as log
import xlrd, xlwt
import numpy as np
import time
import datetime
import scipy.io
import scipy
import pandas as pd
import shutil   
import uuid
import matplotlib.pyplot as plt   
import importlib
import getpass
from copy import deepcopy
from tqdm import tqdm
from scipy.interpolate import interp1d
import pylab
import seaborn as sns

import xlsxwriter
import openpyxl

#os.chdir('C:\\Users\\sklose\\Documents\\ODYM-RECC-Repos\\RECC-Cu-Repo')
import RECC_Paths # Import path file

DFilePath = os.path.join(RECC_Paths.rawdata_pathIAM,'MESSAGE')
ResultsPath= os.path.join(RECC_Paths.rawdata_pathIAM,'Results')
# Define indices for MESSAGE data
Regions_M  = ['AFR','CPA','EEU','FSU','LAC','MEA','NAM','PAO','PAS','SAS','WEU','World']
Model_M    = ['CD_Links_SSP1_v2','CD_Links_SSP2_v2','CD_Links_SSP3_v2']
Scenario_M = ['NPi','NPi2020_1000'] # for BAU/no policy scenario after 2030 and 2dC scenario
Years_M    = [1990,1995,2000,2005,2010,2020,2030,2040,2050,2060,2070,2080,2090,2100,2110]
# For meaning of scenarios, cf. https://db1.ene.iiasa.ac.at/CDLINKSDB/dsd?Action=htmlpage&page=10

# Define indices for ODYM-RECC data
Scenario_R = ['SSP1','SSP2','SSP3']
RCPScen_R  = ['Baseline(unmitigated)','RCP2.6']
Regions_R  = ['R32CAN','R32CHN','R32EU12-M','R32IND','R32JPN','R32USA','France','Germany','Italy','Poland','Spain','UK','Oth_R32EU15','Oth_R32EU12-H','World']
Time_R     = np.arange(2000,2101,1)
TimeL_R    = [i for i in Time_R] 
# Read data into pandas dataframe:
DF = pd.read_csv(os.path.join(DFilePath,'CD_Links_SSP1-3.csv'), sep = ',', encoding = 'unicode_escape')


DF.head(5)

Path_Copper_GW = os.path.join(DFilePath) + '\\Copper_MI.xlsx'   ##Import Data on Material Intensity of kommer [kt/GW]
CopperGW_WB_df = pd.read_excel(Path_Copper_GW)

CopperGW_WB_df.set_index(CopperGW_WB_df['VARIABLE'])



print(DF.shape)
print(len(DF.index))
list(DF.columns.values) # all columns used


       
    
####
# Translate Capacity additions into RECC model dataformat
###





C_EGList_MSG = [##'Capacity|Electricity|Hydro',
                'Capacity|Electricity|Hydro|1',
                'Capacity|Electricity|Hydro|2',
                ##'Capacity|Electricity|Nuclear',
                'Capacity|Electricity|Nuclear|1',
                'Capacity|Electricity|Nuclear|2',
                ##'Capacity|Electricity|Oil',
                ##'Capacity|Electricity|Oil|w/ CCS',
                ##'Capacity|Electricity|Oil|w/o CCS',
                'Capacity|Electricity|Oil|w/o CCS|1',
                'Capacity|Electricity|Oil|w/o CCS|2',
                'Capacity|Electricity|Oil|w/o CCS|3',
                ##'Capacity|Electricity|Solar',
                #'Capacity|Electricity|Solar|CSP',
                'Capacity|Electricity|Solar|CSP|1',
                'Capacity|Electricity|Solar|CSP|2',
                'Capacity|Electricity|Solar|PV',
                ##'Capacity|Electricity|Wind',
                'Capacity|Electricity|Wind|Offshore',
                'Capacity|Electricity|Wind|Onshore',
                #Capacity|Electricity'
                ##'Capacity|Electricity|Biomass',
               # 'Capacity|Electricity|Biomass|w/ CCS',
                'Capacity|Electricity|Biomass|w/ CCS|1',
               # 'Capacity|Electricity|Biomass|w/o CCS',
                'Capacity|Electricity|Biomass|w/o CCS|1',
                'Capacity|Electricity|Biomass|w/o CCS|2',
                ##'Capacity|Electricity|Coal',
            #    'Capacity|Electricity|Coal|w/ CCS',
                'Capacity|Electricity|Coal|w/ CCS|1',
                'Capacity|Electricity|Coal|w/ CCS|2',
               # 'Capacity|Electricity|Coal|w/o CCS',
                'Capacity|Electricity|Coal|w/o CCS|1',
                'Capacity|Electricity|Coal|w/o CCS|2',
                'Capacity|Electricity|Coal|w/o CCS|3',
                'Capacity|Electricity|Coal|w/o CCS|4',
                ##'Capacity|Electricity|Gas',
            #    'Capacity|Electricity|Gas|w/ CCS',
                'Capacity|Electricity|Gas|w/ CCS|1',
             #   'Capacity|Electricity|Gas|w/o CCS',
                'Capacity|Electricity|Gas|w/o CCS|1',
                'Capacity|Electricity|Gas|w/o CCS|2',
                'Capacity|Electricity|Gas|w/o CCS|3',
                'Capacity|Electricity|Geothermal'
                ]



CA_EGList_MSG = [##'Capacity Additions|Electricity|Hydro',
            'Capacity Additions|Electricity|Hydro|1',
            'Capacity Additions|Electricity|Hydro|2',
            ##'Capacity Additions|Electricity|Nuclear',
            'Capacity Additions|Electricity|Nuclear|1',
            'Capacity Additions|Electricity|Nuclear|2',
            ##'Capacity Additions|Electricity|Oil',
            #'Capacity Additions|Electricity|Oil|w/ CCS',
            #'Capacity Additions|Electricity|Oil|w/o CCS',
            'Capacity Additions|Electricity|Oil|w/o CCS|1',
            'Capacity Additions|Electricity|Oil|w/o CCS|2',
            'Capacity Additions|Electricity|Oil|w/o CCS|3',
            ##'Capacity Additions|Electricity|Solar',
            #'Capacity Additions|Electricity|Solar|CSP',
            'Capacity Additions|Electricity|Solar|CSP|1',
            'Capacity Additions|Electricity|Solar|CSP|2',
            'Capacity Additions|Electricity|Solar|PV',
            ##'Capacity Additions|Electricity|Wind',
            'Capacity Additions|Electricity|Wind|Offshore',
            'Capacity Additions|Electricity|Wind|Onshore',
           # 'Capacity Additions|Electricity',
            ##'Capacity Additions|Electricity|Biomass',
           # 'Capacity Additions|Electricity|Biomass|w/ CCS',
            'Capacity Additions|Electricity|Biomass|w/ CCS|1',
            #'Capacity Additions|Electricity|Biomass|w/o CCS',
            'Capacity Additions|Electricity|Biomass|w/o CCS|1',
            'Capacity Additions|Electricity|Biomass|w/o CCS|2',
            ##'Capacity Additions|Electricity|Coal',
         #   'Capacity Additions|Electricity|Coal|w/ CCS',
            'Capacity Additions|Electricity|Coal|w/ CCS|1',
            'Capacity Additions|Electricity|Coal|w/ CCS|2',
          #  'Capacity Additions|Electricity|Coal|w/o CCS',
            'Capacity Additions|Electricity|Coal|w/o CCS|1',
            'Capacity Additions|Electricity|Coal|w/o CCS|2',
            'Capacity Additions|Electricity|Coal|w/o CCS|3',
            'Capacity Additions|Electricity|Coal|w/o CCS|4',
            ##'Capacity Additions|Electricity|Gas',
        #    'Capacity Additions|Electricity|Gas|w/ CCS',
            'Capacity Additions|Electricity|Gas|w/ CCS|1',
        #    'Capacity Additions|Electricity|Gas|w/o CCS',
            'Capacity Additions|Electricity|Gas|w/o CCS|1',
            'Capacity Additions|Electricity|Gas|w/o CCS|2',
            'Capacity Additions|Electricity|Gas|w/o CCS|3',
            'Capacity Additions|Electricity|Geothermal'
            ]


CA_List_ODYM = ['Capacity additions solar photovoltaic power plant',
            'Capacity additions concentrating solar power plant (CSP)',
            #'concentrating solar power plant (CSP) with solar multiple of 3',
            'Capacity additions wind power plant onshore',
            'Capacity additions wind power plant offshore',
            'Capacity additions hydro power plant',
            'Capacity additions nuclear power plant', 
            'Capacity additions coal power plant',
           # 'coal power plant without abatement measures',
            'Capacity additions bio powerplant',
            'Capacity additions oil power plant',
            'Capacity additions geothermal power plant',
           # 'IGCC power plant',
         #  'light oil combined cycle',
            'Capacity additions gas combined cycle power plant',
            'Capacity additions advanced coal power plant with CCS',
            'Capacity additions IGCC power plant with CCS',
            'Capacity additions biomass power plant with CCS',
            'Capacity additions gas combined cycle power plant with CCS',
          #  'storage technologies'
            ]

C_List_ODYM = ['Capacity solar photovoltaic power plant',
            'Capacity concentrating solar power plant (CSP)',
            #'concentrating solar power plant (CSP) with solar multiple of 3',
            'Capacity wind power plant onshore',
            'Capacity wind power plant offshore',
            'Capacity hydro power plant',
            'Capacity nuclear power plant', 
            'Capacity coal power plant',
           # 'coal power plant without abatement measures',
            'Capacity bio powerplant',
            'Capacity oil power plant',
            'Capacity geothermal power plant',
           # 'IGCC power plant',
         #  'light oil combined cycle',
            'Capacity gas combined cycle power plant',
            'Capacity advanced coal power plant with CCS',
            'Capacity coal power plant with CCS',
            'Capacity biomass power plant with CCS',
            'Capacity gas combined cycle power plant with CCS',
          #  'storage technologies'
            ]


Lft_EGList_MSG = [#'Lifetime|Electricity|Hydro',
            'Lifetime|Electricity|Hydro|1',
            'Lifetime|Electricity|Hydro|2',
            #'Lifetime|Electricity|Nuclear',
            'Lifetime|Electricity|Nuclear|1',
            'Lifetime|Electricity|Nuclear|2',
           # 'Lifetime|Electricity|Oil|w/o CCS',
            'Lifetime|Electricity|Oil|w/o CCS|1',
            'Lifetime|Electricity|Oil|w/o CCS|2',
            'Lifetime|Electricity|Oil|w/o CCS|3',
            'Lifetime|Electricity|Solar|CSP|1',
            'Lifetime|Electricity|Solar|CSP|2',
            'Lifetime|Electricity|Solar|PV',
            'Lifetime|Electricity|Wind|Offshore',
            'Lifetime|Electricity|Wind|Onshore',
            'Lifetime|Electricity|Biomass|w/ CCS|1',
            'Lifetime|Electricity|Biomass|w/o CCS|1',
            'Lifetime|Electricity|Biomass|w/o CCS|2',
            'Lifetime|Electricity|Coal|w/ CCS|1',
            'Lifetime|Electricity|Coal|w/ CCS|2',
            'Lifetime|Electricity|Coal|w/o CCS|1',
            'Lifetime|Electricity|Coal|w/o CCS|2',
            'Lifetime|Electricity|Coal|w/o CCS|3',
            'Lifetime|Electricity|Coal|w/o CCS|4',
            'Lifetime|Electricity|Gas|w/ CCS|1',
            'Lifetime|Electricity|Gas|w/o CCS|1',
            'Lifetime|Electricity|Gas|w/o CCS|2',
            'Lifetime|Electricity|Gas|w/o CCS|3',
            'Lifetime|Electricity|Geothermal'
]


EG_List_ODYM = ['solar photovoltaic power plant',
            'concentrating solar power plant (CSP)',
            #'concentrating solar power plant (CSP) with solar multiple of 3',
            'wind power plant onshore',
            'wind power plant offshore',
            'hydro power plant',
            'nuclear power plant', 
            'coal power plant',
           # 'coal power plant without abatement measures',
            'bio powerplant',
            'oil power plant',
            'geothermal power plant',
           # 'IGCC power plant',
         #  'light oil combined cycle',
            'gas combined cycle power plant',
            'advanced coal power plant with CCS',
            'coal power plant with CCS',
            'biomass power plant with CCS',
            'gas combined cycle power plant with CCS',
          #  'storage technologies'
            ]


Lft_List_ODYM = ['Lifetime solar photovoltaic power plant',
            'Lifetime concentrating solar power plant (CSP)',
            #'concentrating solar power plant (CSP) with solar multiple of 3',
            'Lifetime wind power plant onshore',
            'Lifetime wind power plant offshore',
            'Lifetime hydro power plant',
            'Lifetime nuclear power plant', 
            'Lifetime coal power plant',
           # 'coal power plant without abatement measures',
            'Lifetime bio powerplant',
            'Lifetime oil power plant',
            'Lifetime geothermal power plant',
           # 'IGCC power plant',
         #  'light oil combined cycle',
            'Lifetime gas combined cycle power plant',
            'Lifetime advanced coal power plant with CCS',
            'Lifetime coal power plant with CCS',
            'Lifetime biomass power plant with CCS',
            'Lifetime gas combined cycle power plant with CCS',
          #  'storage technologies'
            ]


EGListLookup = [4,
                4,
                5,
                5,
                8,
                8,
                8,
                1,
                1,
                0,
                3,
                2,
                13,
                7,
                7,
                12,
                12,
                6,
                6,
                6,
                6,
                14,
                10,
                10,
                10,
                9]





#### Convert MESSAGE Cappacity Addition data in ODYM format

CA= DF.loc[DF['VARIABLE'].isin(CA_EGList_MSG) & DF['SCENARIO'].isin(Scenario_M) & DF['MODEL'].isin(Model_M)] # extract emissions from energy supply, in Mt CO2/yr
        
[CA['VARIABLE'].replace(   to_replace=[CA_EGList_MSG[i]],value=EG_List_ODYM[EGListLookup[i]],   inplace=True) for i in range(0,len(CA_EGList_MSG))]    
[CA['MODEL'].replace(   to_replace=[Model_M[i]],value=Scenario_R[i],   inplace=True) for i in range(0,3)]
[CA['SCENARIO'].replace(to_replace=[Scenario_M[i]],value=RCPScen_R[i], inplace=True) for i in range(0,2)]
        
### sum over ALL values with same model, scenario and Variable
dCA = CA.groupby(['REGION', 'MODEL','SCENARIO','VARIABLE']).sum()

ddCA = dCA.loc[:, (dCA != 0).any(axis=0)] ##delete columns that are zero





#### Convert MESSAGE Cappacity (Stock) data in ODYM format

Capacity= DF.loc[DF['VARIABLE'].isin(C_EGList_MSG) & DF['SCENARIO'].isin(Scenario_M) & DF['MODEL'].isin(Model_M)] # extract emissions from energy supply, in Mt CO2/yr
        
[Capacity['VARIABLE'].replace(   to_replace=[C_EGList_MSG[i]],value=EG_List_ODYM[EGListLookup[i]],   inplace=True) for i in range(0,len(C_EGList_MSG))]    
[Capacity['MODEL'].replace(   to_replace=[Model_M[i]],value=Scenario_R[i],   inplace=True) for i in range(0,3)]
[Capacity['SCENARIO'].replace(to_replace=[Scenario_M[i]],value=RCPScen_R[i], inplace=True) for i in range(0,2)]
        
### sum over ALL values with same model, scenario and Variable
dCapacity = Capacity.groupby(['REGION', 'MODEL','SCENARIO','VARIABLE']).sum()

ddCapacity = dCapacity.loc[:, (dCapacity != 0).any(axis=0)] ##delete columns that are zero



### Merging List of Products with Lifetime/Copper intensity List
ddCA2 = ddCA.reset_index(level=['REGION', 'MODEL','SCENARIO','VARIABLE'])
DF_Merge2 = ddCA2.merge(CopperGW_WB_df, on = ['VARIABLE'])
DF_Merge = DF_Merge2.set_index(['REGION', 'MODEL','SCENARIO','VARIABLE'])   # Set Indexes again 


df_CopperIntensity = DF_Merge[CopperGW_WB_df.columns.values[1]] #Matched copper intensities to Variables in Dataframe


#####Create stepfunction and expand input file to every year input
col_1990 = np.arange(1986,1990)
col_1990j = " ".join(str(x) for x in col_1990)
col_1990s = col_1990j.split()

col_1995 = np.arange(1991,1995)
col_1995j = " ".join(str(x) for x in col_1995)
col_1995s = col_1995j.split()

col_2000 = np.arange(1996,2000)
col_2000j = " ".join(str(x) for x in col_2000)
col_2000s = col_2000j.split()  

col_2005 = np.arange(2001,2005)
col_2005j = " ".join(str(x) for x in col_2005)
col_2005s = col_2005j.split()  

col_2010 = np.arange(2006,2010)
col_2010j = " ".join(str(x) for x in col_2010)
col_2010s = col_2010j.split()  

col_2020 = np.arange(2011,2020)
col_2020j = " ".join(str(x) for x in col_2020)
col_2020s = col_2020j.split()  

col_2030 = np.arange(2021,2030)
col_2030j = " ".join(str(x) for x in col_2030)
col_2030s = col_2030j.split()  

col_2040 = np.arange(2031,2040)
col_2040j = " ".join(str(x) for x in col_2040)
col_2040s = col_2040j.split()  

col_2050 = np.arange(2041,2050)
col_2050j = " ".join(str(x) for x in col_2050)
col_2050s = col_2050j.split()  

col_2060 = np.arange(2051,2060)
col_2060j = " ".join(str(x) for x in col_2060)
col_2060s = col_2060j.split()  

col_2070 = np.arange(2061,2070)
col_2070j = " ".join(str(x) for x in col_2070)
col_2070s = col_2070j.split()  

col_2080 = np.arange(2071,2080)
col_2080j = " ".join(str(x) for x in col_2080)
col_2080s = col_2080j.split()  

col_2090 = np.arange(2081,2090)
col_2090j = " ".join(str(x) for x in col_2090)
col_2090s = col_2090j.split()  

col_2100 = np.arange(2091,2100)
col_2100j = " ".join(str(x) for x in col_2100)
col_2100s = col_2100j.split()  

col_2110 = np.arange(2101,2110)
col_2110j = " ".join(str(x) for x in col_2110)
col_2110s = col_2110j.split()  

ddCAexp0 = pd.concat([ddCA, pd.DataFrame(columns = col_1990s )],sort=True)
ddCAexp1 = pd.concat([ddCAexp0, pd.DataFrame(columns = col_1995s )],sort=True)
ddCAexp2 = pd.concat([ddCAexp1, pd.DataFrame(columns = col_2000s )],sort=True)
ddCAexp3 = pd.concat([ddCAexp2, pd.DataFrame(columns = col_2005s )],sort=True)
ddCAexp4 = pd.concat([ddCAexp3, pd.DataFrame(columns = col_2010s )],sort=True)
ddCAexp5 = pd.concat([ddCAexp4, pd.DataFrame(columns = col_2020s )],sort=True)
ddCAexp6 = pd.concat([ddCAexp5, pd.DataFrame(columns = col_2030s )],sort=True)
ddCAexp7 = pd.concat([ddCAexp6, pd.DataFrame(columns = col_2040s )],sort=True)
ddCAexp8 = pd.concat([ddCAexp7, pd.DataFrame(columns = col_2050s )],sort=True)
ddCAexp9 = pd.concat([ddCAexp8, pd.DataFrame(columns = col_2060s )],sort=True)
ddCAexp10 = pd.concat([ddCAexp9, pd.DataFrame(columns = col_2070s )],sort=True)
ddCAexp11 = pd.concat([ddCAexp10, pd.DataFrame(columns = col_2080s )],sort=True)
ddCAexp12 = pd.concat([ddCAexp11, pd.DataFrame(columns = col_2090s )],sort=True)
ddCAexp13 = pd.concat([ddCAexp12, pd.DataFrame(columns = col_2100s )],sort=True)
ddCAexp14 = pd.concat([ddCAexp13, pd.DataFrame(columns = col_2100s )],sort=True)
ddCAexp15 = pd.concat([ddCAexp14, pd.DataFrame(columns = col_2110s )],sort=True)

for c in col_1990s:
    ddCAexp15[c] = ddCAexp15['1990']
for c in col_1995s:
    ddCAexp15[c] = ddCAexp15['1995']
for c in col_2000s:
    ddCAexp15[c] = ddCAexp15['2000']
for c in col_2005s:
    ddCAexp15[c] = ddCAexp15['2005']
for c in col_2010s:
    ddCAexp15[c] = ddCAexp15['2010']
for c in col_2020s:
    ddCAexp15[c] = ddCAexp15['2020']
for c in col_2030s:
    ddCAexp15[c] = ddCAexp15['2030']
for c in col_2040s:
    ddCAexp15[c] = ddCAexp15['2040']
for c in col_2050s:
    ddCAexp15[c] = ddCAexp15['2050']
for c in col_2060s:
    ddCAexp15[c] = ddCAexp15['2060']
for c in col_2070s:
    ddCAexp15[c] = ddCAexp15['2070']
for c in col_2080s:
    ddCAexp15[c] = ddCAexp15['2080']
for c in col_2090s:
    ddCAexp15[c] = ddCAexp15['2090']
for c in col_2100s:
    ddCAexp15[c] = ddCAexp15['2100']
for c in col_2110s:
    ddCAexp15[c] = ddCAexp15['2110']

## Calculated intensities and interpolate
#Idx_Time_Rel = [i -1990 for i in Years_M]
#for mn in range(0,len(ddCA)):
#    f_DF = interp1d(Idx_Time_Rel, np.array(ddCA.iloc[mn][0:15]), kind = 'next')
#    f_DF_CAint[mn,:]   = f_DF(np.arange(0,121))
#               # GHG_Intp[mn,r,S,R,:]   = f_GHG(np.arange(0,101))
#             #   Enrgy_Intp[mn,r,S,R,:] = f_Egy(np.arange(0,101))
#
#exp_int = pd.DataFrame(f_DF_CAint,index=ddCA.index)
#
#exp_merged_indexaligned = exp_int.reset_index(level=['REGION', 'MODEL','SCENARIO','VARIABLE'])
#
#DF_Merge2_int = exp_merged_indexaligned.merge(CopperGW_WB_df, on = ['VARIABLE'])
#
#
#DF_Merge2_int.columns.values[4:125]=DF_Merge2_int.columns.values[4:125]+1990
#
##              
##        
#


#### Save future Service demand in RECC format

col_future_demand = np.arange(2014,2110)
col_future_demandj = " ".join(str(x) for x in col_future_demand)
col_future_demands = col_future_demandj.split()
   
df_future_demand = ddCAexp15[col_future_demands]   



df_future_demandI = df_future_demand.reset_index(level=['REGION', 'MODEL','SCENARIO','VARIABLE'])


future_demand = pd.DataFrame(df_future_demandI)

book=openpyxl.load_workbook(DFilePath+'\\2_S_RECC_FinalProducts_Future_EGT_V1.0.xlsx' )
        
writer=pd.ExcelWriter(DFilePath+'\\2_S_RECC_FinalProducts_Future_EGT_V1.0.xlsx', engine='openpyxl')

writer.book=book

writer.sheets = dict((ws.title, ws) for ws in book.worksheets)


future_demand.to_excel(writer, sheet_name='Sheet1',startcol=0,startrow=0,index=False,header=True)


writer.save()


### Safe Copper Intensity dataframe RECC format

df_CopperIntensity = df_CopperIntensity.reset_index(level=['REGION', 'MODEL','SCENARIO','VARIABLE'])

CopperIntensity = pd.DataFrame(df_CopperIntensity)

book=openpyxl.load_workbook(DFilePath+'\\3_MC_RECC_EGT_V1.0.xlsx' )
        
writer=pd.ExcelWriter(DFilePath+'\\3_MC_RECC_EGT_V1.0.xlsx', engine='openpyxl')

writer.book=book

writer.sheets = dict((ws.title, ws) for ws in book.worksheets)


CopperIntensity.to_excel(writer, sheet_name='values',startcol=0,startrow=0,index=False,header=True)


writer.save()




#### Save Lifetime Dataframe in RECC format

Lifetime= DF.loc[DF['VARIABLE'].isin(Lft_EGList_MSG) & DF['SCENARIO'].isin(Scenario_M) & DF['MODEL'].isin(Model_M)] # extract emissions from energy supply, in Mt CO2/yr

[Lifetime['VARIABLE'].replace(   to_replace=[Lft_EGList_MSG[i]],value=EG_List_ODYM[EGListLookup[i]],   inplace=True) for i in range(0,len(Lft_EGList_MSG))]    
[Lifetime['MODEL'].replace(   to_replace=[Model_M[i]],value=Scenario_R[i],   inplace=True) for i in range(0,3)]
[Lifetime['SCENARIO'].replace(to_replace=[Scenario_M[i]],value=RCPScen_R[i], inplace=True) for i in range(0,2)]

dLifetime = Lifetime.groupby(['REGION', 'MODEL','SCENARIO','VARIABLE']).max()

dLifetime['copper electric grade'] = dLifetime.max(axis=1)

RECC_Lifetime= pd.DataFrame(dLifetime['copper electric grade'])


RECC_LifetimeI = RECC_Lifetime.reset_index(level=['REGION', 'MODEL','SCENARIO','VARIABLE'])


book=openpyxl.load_workbook(DFilePath+'\\3_LT_RECC_ProductLifetime_EGT_V1.0.xlsx' )
        
writer=pd.ExcelWriter(DFilePath+'\\3_LT_RECC_ProductLifetime_EGT_V1.0.xlsx', engine='openpyxl')

writer.book=book

writer.sheets = dict((ws.title, ws) for ws in book.worksheets)


RECC_LifetimeI.to_excel(writer, sheet_name='values',startcol=0,startrow=0,index=False,header=True)


writer.save()



#### Save final products 2015 in RECC format


col_final_products_2015 = np.arange(1990,2015)
col_final_products_2015j = " ".join(str(x) for x in col_final_products_2015)
col_final_products_2015s = col_final_products_2015j.split()


df_final_products_2015 = ddCAexp15[col_final_products_2015s] 

df_final_products_2015['0']=2015

#Unstack= ddCAexp15.unstack(level=[3])
#Unstack2= Unstack.unstack(level=[2])
#Stack= ddCAexp15.stack(level=-1, dropna=True)
df_final_products_20152 = df_final_products_2015.reset_index(level=['REGION', 'MODEL','SCENARIO','VARIABLE'])


df_final_products_2015_merged = df_final_products_20152.merge(RECC_Lifetime, on = ['REGION', 'MODEL','SCENARIO','VARIABLE'])




book=openpyxl.load_workbook(DFilePath+'\\2_S_RECC_FinalProducts_2015_EGT_V1.0.xlsx' )
        
writer=pd.ExcelWriter(DFilePath+'\\2_S_RECC_FinalProducts_2015_EGT_V1.0.xlsx', engine='openpyxl')

writer.book=book

writer.sheets = dict((ws.title, ws) for ws in book.worksheets)


df_final_products_2015_merged.to_excel(writer, sheet_name='values',startcol=0,startrow=0,index=False,header=True)


writer.save()
# The end.

